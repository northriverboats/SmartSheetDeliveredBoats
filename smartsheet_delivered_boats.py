#!/usr/bin/env python3

import click
import datetime
import glob
import logging
import openpyxl
import os
import smartsheet
import subprocess
import sys
from dotenv import load_dotenv
from emailer import mail_results
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

reports = [
    {'id': '7295282737112964', 'name': 'Alaska Frontier Fabrication - Delivered Boats'},
    {'id': '2511118666098564', 'name': 'Avataa - Delivered Boats'},
    {'id': '8402628385236868', 'name': 'Boat Country - Delivered Boats'},
    {'id': '1084278990759812', 'name': 'Clemens Eugene - Delivered Boats'},
    {'id': '381227204142980',  'name': 'Clemens Portland - Delivered Boats'},
    {'id': '5861565896386436', 'name': 'Drummondville Marine - Delivered Boats'},
    {'id': '8121565725386628', 'name': 'Elephant Boys - Delivered Boats'},
    {'id': '3613345434625924', 'name': 'Ennis Brothers - Delivered Boats'},
    {'id': '8947439265703812', 'name': 'Erie Marine Sales LLC - Delivered Boats'},
    {'id': '4252384307242884', 'name': 'Idaho Marine - Delivered Boats'},
    {'id': '6960979662661508', 'name': 'PGM - Delivered Boats'},
    {'id': '6995614278936452', 'name': 'Port Boat House - Delivered Boats'},
    {'id': '7841138720696196', 'name': 'RF Marina - Delivered Boats'},
    {'id': '8614507711883140', 'name': 'The Bay Co - Delivered Boats'},
    {'id': '1332356301776772', 'name': 'Three Rivers - Delivered Boats'},
    {'id': '3513735831676804', 'name': 'Valley Marine - Delivered Boats'},
    {'id': '8017335459047300', 'name': 'Y Marina - Delivered Boats'},
]

log_text = ""
errors = False

def log(text, error=None):
    global log_text, errors
    print(text)
    log_text += text + "\n"
    if (error):
        errors = True


def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)


def normal_border(wsNew, row):
    for i in range(1,11):
        side1 = 'thin'
        side2 = 'thin'
        if i == 10:
            side1 = 'medium'
        if i == 1:
            side2 = 'medium'
        wsNew.cell(column=i,row=row+7).border = Border(right = Side(border_style=side1, color='FF000000'),
                                                       left = Side(border_style=side2, color='FF000000'))

def heading_border(wsNew, row):
    for i in range(1,11):
        side1 = 'thin'
        side2 = 'thin'
        if i == 10:
            side1 = 'medium'
        if i == 1:
            side2 = 'medium'
        wsNew.cell(column=i,row=row+7).border = Border(right = Side(border_style=side1, color='FF000000'),
                                                       left = Side(border_style=side2, color='FF000000'),
                                                       top = Side(border_style='medium', color='FF000000'),
                                                       bottom = Side(border_style='medium', color='FF000000'))

def end_page_border(wsNew, row):
    for i in range(1,11):
        side1 = 'thin'
        side2 = 'thin'
        if i == 10:
            side1 = 'medium'
        if i == 1:
            side2 = 'medium'
        wsNew.cell(column=i,row=row+7).border = Border(right = Side(border_style=side1, color='FF000000'),
                                                       left = Side(border_style=side2, color='FF000000'),
                                                       bottom = Side(border_style='medium', color='FF000000'))

def bottom_border(wsNew, row):
    for i in range(1,11):
        side1 = 'thin'
        side2 = 'thin'
        if i == 10:
            side1 = 'medium'
        if i == 1:
            side2 = 'medium'
        wsNew.cell(column=i,row=row+7).border = Border(right = Side(border_style=side1, color='FF000000'),
                                                       left = Side(border_style=side2, color='FF000000'),
                                                       bottom = Side(border_style='medium', color='FF000000'))

def fetch_value(cell):
    value = cell.value
    if cell.data_type == 's':
        return value
    if cell.is_date:
        return ('%02d/%02d/%02d' %(value.month,value.day,value.year-2000))
    if value == None:
        return ''
    return str(int(value))

def set_mast_header(wsNew, logo_name, dealer_name):
    # place logo and dealername on new sheet
    img = Image(logo_name)
    wsNew.add_image(img, 'B1')
    wsNew['B5'] = dealer_name
    wsNew['J5'] = "Report Date: %s " % (datetime.datetime.today().strftime('%m/%d/%Y'))

def set_header(wsNew, row):
    titles = ['Hull #',
              'Boat Model',
              'Order Details',
              'Colors Interior / Exterior',
              'Engines',
              'Current Phase',
              'Est Start/Finish',
              'Actual Start',
              'Actual Finish',
              'Notes'
             ]
    heading_border(wsNew,row)
    wsNew.row_dimensions[row+7].height = 21.6
    for i in range(1,11):
        wsNew.cell(row=row+7, column=i,value=titles[i-1])
        wsNew.cell(row=row+7, column=i).alignment = Alignment(horizontal='center',vertical='center')


def set_footer(wsNew, row):
    normal_border(wsNew,row)
    normal_border(wsNew,row+1)

    wsNew.merge_cells(start_row=row+8, start_column=1, end_row=row+8, end_column=3)
    wsNew.cell(row=row+8, column=1,value="Contact Joe for 9'6 build dates")
    wsNew.cell(row=row+8, column=1).alignment = Alignment(horizontal='center')
    wsNew.cell(row=row+8, column=1).font = Font(bold=True)

    wsNew.merge_cells(start_row=row+9, start_column=1, end_row=row+9, end_column=10)
    wsNew.cell(row=row+9, column=1,value="NOTE: Estimated Start & Delivery Week's can be 1 - 2 Weeks before or after original dates")
    wsNew.cell(row=row+9, column=1).alignment = Alignment(horizontal='center')
    wsNew.cell(row=row+9, column=1).font = Font(bold=True)
    bottom_border(wsNew,row+2)


def process_row(wsOld,wsNew,row,offset,bgColor,base): #base 7 or base 0
    for i in range(1,11):
        value = fetch_value(wsOld.cell(column=i,row=row))
        cell = wsNew.cell(column=i,row=row+base+offset)
        cell.value = value
        bg = bgColor
        if i == 3 and cell.value.lower().find('stock') == -1:
            bg = 'FFFFC000'
        if wsOld.cell(column=i,row=row).fill.start_color.index == 'FF00CA0E':
            bg = 'FF00CA0E'
        if bg != None:
            cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type = "solid")
        if i == 2 or i == 4 or i == 5 or i == 10:
            cell.font = Font(size=8)

def process_rows(wsOld,wsNew,base,forPDF):
    pagelen = 60
    offset = 0
    for i in range(2,wsOld.max_row+1):
        if i ==  pagelen + 3:
            pagelen = 61
        if forPDF:
            if (i+base+1) % pagelen == 0:
                set_header(wsNew,i+offset)
                offset = offset + 1
            if ((wsOld.max_row +base) % pagelen == pagelen-2) and  (wsOld.max_row == i):
                set_header(wsNew,i+offset -1)
            if ((wsOld.max_row +base) % pagelen == pagelen-3) and  (wsOld.max_row-1 == i):
                offset = offset + 3
                set_header(wsNew,i+offset -1)
            if ((wsOld.max_row +base) % pagelen == pagelen-4) and  (wsOld.max_row-2 == i):
                offset = offset + 5
                set_header(wsNew,i+offset -1)

        bgColor = None
        model = wsOld["B"+str(i)].value

        if model.find("OS") != -1:
            bgColor = 'FFA6A6A6'
        if model.replace(" ","").lower().find('hardtop') != -1:
            bgColor = 'FFD9D9D9'
        process_row(wsOld,wsNew,i,offset,bgColor,base)

        if (i+base+1) % pagelen == pagelen-1 and wsOld.max_row != i and forPDF:
            end_page_border(wsNew,i+offset)
        elif ((wsOld.max_row + base) % pagelen == pagelen-2) and  (wsOld.max_row-1 == i) and forPDF:
            end_page_border(wsNew,i+offset)
            offset = offset + 1
        elif ((wsOld.max_row + base) % pagelen == pagelen-3) and  (wsOld.max_row-2 == i) and forPDF:
            end_page_border(wsNew,i+offset)
        elif ((wsOld.max_row + base) % pagelen == pagelen-4) and  (wsOld.max_row-3 == i) and forPDF:
            end_page_border(wsNew,i+offset)
        else:
            normal_border(wsNew,i+offset)
    set_footer(wsNew, wsOld.max_row+offset+1)
    return offset

def process_sheet_to_pdf(source_dir, target_dir, file):
    # change variables here
    input_name = source_dir + 'downloads/' + file
    pdf_dir = (target_dir + 'Formatted - PDF/')
    output_name = pdf_dir + file
    logo_name = source_dir + 'nrblogo1.jpg'
    dealer_name = input_name[42:-22]
    base = 7

    # load sheet data is coming from
    wbOld = openpyxl.load_workbook(input_name)
    wsOld = wbOld.active

    # load sheet we are copying data to
    wbNew = openpyxl.load_workbook(source_dir + 'DeliveredBoatsTemplate.xlsx')
    wsNew = wbNew.active

    set_mast_header(wsNew, logo_name, dealer_name)
    offset = process_rows(wsOld, wsNew, base, True)

    range = 'A1:J'+str(wsOld.max_row+base+offset+3)

    wbNew.create_named_range('_xlnm.Print_Area', wsNew, range, scope=0)

    # save new sheet out to new file
    try:
        wbNew.save(output_name)
        result = subprocess.call(['/usr/bin/unoconv',
                         '-f', 'pdf',
                         '-t', source_dir + 'landscape.ots',
                         '--output='+ output_name[:-3] + 'pdf',
                         output_name])
        if (result):
            log('             UNICONV FAILED TO CREATE PDF', True)
    except Exception as e:
        log('             FAILED TO CREATE XLSX AND PDF: ' + str(e), True)


def process_sheet_to_xlsx(source_dir, target_dir, file):
    # change variables here
    input_name = source_dir + 'downloads/' + file
    output_name = target_dir + file
    logo_name = source_dir + 'nrblogo1.jpg'
    dealer_name = file[:-22]
    base = 7

    # load sheet data is coming from
    wbOld = openpyxl.load_workbook(input_name)
    wsOld = wbOld.active

    # load sheet we are copying data to
    wbNew = openpyxl.load_workbook(source_dir + 'DeliveredBoatsTemplate.xlsx')
    wsNew = wbNew.active
    set_mast_header(wsNew, logo_name, dealer_name)
    offset = process_rows(wsOld, wsNew, base, False)
    range = 'A1:J'+str(wsOld.max_row+3)
    wbNew.create_named_range('_xlnm.Print_Area', wsNew, range, scope=0)

    # save new sheet out to new file
    try:
        wbNew.save(output_name)
    except Exception as e:
        log('             FAILED TO CREATE XLSX: ' + str(e), True)


def process_sheets(source_dir, target_dir, pdf, excel):
    log("\nPROCESS SHEETS ===============================")
    os.chdir(source_dir + 'downloads/')
    for file in sorted(glob.glob('*.xlsx')):
        if pdf:
          log("  converting %s to pdf" % (file))
          process_sheet_to_pdf(source_dir, target_dir, file)
        if excel:
          log("  converting %s to xlsx" % (file))
          process_sheet_to_xlsx(source_dir, target_dir, file)
        log("")


def download_sheets(dealers, api, source_dir):
    files = os.listdir(source_dir + 'downloads')
    for file in files:
        os.remove(os.path.join(source_dir + 'downloads', file))

    smart = smartsheet.Smartsheet(api)
    smart.assume_user(os.getenv('SMARTSHEET_USER'))
    log("DOWNLOADING SHEETS ===========================")
    for report in dealers:
        log("  downloading sheet: " + report['name'])
        try:
            smart.Reports.get_report_as_excel(report['id'], source_dir + 'downloads')
        except Exception as e:
            log('                     ERROR DOWNLOADING SHEET: ' + str(e), True)


def send_error_report():
    subject = 'Smartsheet Deliverd Boats Error Report'
    mail_results(subject, log_text)


@click.command()
@click.option(
    '--list',
    '-l',
    'list_',
    is_flag=True,
    help='Print list of dealers'
)
@click.option(
    '--dealer',
    '-d',
    multiple=True,
    help='Dealer to include (can use multiple times)'
)
@click.option(
    '--ignore',
    '-i',
    multiple=True,
    help='Dealer to ignore (can use multiple times)'
)
@click.option(
    '--pdf/--no-pdf',
    default=True,
    help='Create PDFs unless --no-pdf'
)
@click.option(
    '--excel/--no-excel',
    default=True,
    help='Create Excel Sheets unless --no-excel'
)
def main(list_, dealer, ignore, pdf, excel):
    # load environmental variables
    env_path = resource_path('.env')
    load_dotenv(dotenv_path=env_path)

    api = os.getenv('SMARTSHEET_API')
    source_dir = os.getenv('SOURCE_DIR')
    target_dir = os.getenv('TARGET_DIR')

    if list_:
        for report in reports:
            print("'" + report['name'].split('-')[0].strip() + "'")
        sys.exit(0)

    # Add dealers we want to report on
    if dealer:
        dealers = [r for r in reports if r['name'].split('-')[0].strip() in dealer]
    else:
        dealers = reports

    # Delete dealers we are not intested in
    if ignore:
        dealers = [d for d in dealers if d['name'].split('-')[0].strip() not in ignore]

    # actual processing
    try:
        download_sheets(dealers, api, source_dir)
        process_sheets(source_dir, target_dir, pdf, excel)
    except Exception as e:
        log('Uncaught Error in main(): ' + str(e), True)
    if (errors):
        send_error_report()
    sys.exit(0)

if __name__ == "__main__":
    main()
